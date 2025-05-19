# Standard library imports
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import tkinter.font as tkfont
from functools import partial
import json

# Third-party imports
from PIL import Image, ImageTk, ImageOps
import requests
import openpyxl
import fitz
from openpyxl.utils import get_column_letter

# Constants
API_KEY = 'ZlA0wiVWdf8AvPodV4MCrXIxs3rcIDAx'
OCR_URL = "https://api.iapp.co.th/ocr/v3/receipt/file"
FONT_FAMILY = "TH Niramit AS"
DEFAULT_FONT_SIZE = 16
PREVIEW_WIDTH = 350  # ขนาดความกว้างของ preview (ลดลง)
PREVIEW_HEIGHT = 450  # ขนาดความสูงของ preview (ลดลง)

class OCRApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ACCESS OCR")
        self.root.geometry("1200x800")
        self.root.configure(bg="#e9ecef")

        # --- สำหรับ multi-image ---
        self.image_paths = []
        self.current_image_index = 0

        self._setup_styles()
        self._setup_main_frame()

        # --- สร้าง left_panel รวม preview + ปุ่ม ---
        left_panel = tk.Frame(self.main_frame, bg="#e9ecef")
        left_panel.grid(row=0, column=0, rowspan=2, padx=(20,10), pady=(20,10), sticky="ns")

        # สร้าง preview_frame ใน left_panel และ pack
        self.preview_frame = tk.Frame(left_panel, bg="#fff", height=PREVIEW_HEIGHT+60, width=PREVIEW_WIDTH+20)
        self.preview_frame.pack(side="top", fill="x", padx=0, pady=0)

        self._setup_preview_frame()

        # --- ปุ่ม Prev/Next ใต้ preview ---
        nav_frame = tk.Frame(self.preview_frame, bg="#fff")
        nav_frame.pack(side="bottom", pady=(2, 0))
        self.prev_btn = ttk.Button(nav_frame, text="⟨ Prev", width=8, command=self.show_prev_image)
        self.prev_btn.pack(side="left", padx=4)
        self.next_btn = ttk.Button(nav_frame, text="Next ⟩", width=8, command=self.show_next_image)
        self.next_btn.pack(side="left", padx=4)
        self.update_nav_buttons()

        # วางปุ่มไว้ด้านล่าง left_panel
        btns_frame_left = tk.Frame(left_panel, bg="#e9ecef")
        btns_frame_left.pack(side="bottom", fill="x", pady=(10,0))
        button_width = 22
        ttk.Button(btns_frame_left, text="Upload File", command=self.select_file, width=button_width).pack(fill="x", pady=(0,6))
        ttk.Button(btns_frame_left, text="ประมวลผล OCR", command=self.process_ocr, width=button_width).pack(fill="x", pady=(0,6))
        ttk.Button(btns_frame_left, text="Export เป็น Excel", command=self.export_excel, width=button_width).pack(fill="x")

        # --- สำหรับ Zoom ---
        self.original_img = None
        self.current_img = None
        self.zoom_ratio = 1.0

        # --- โหลด config (fields และ product_codes) จาก JSON เดียวกัน ---
        config = self.load_config()
        self.fields = config.get("fields", [])
        self.product_codes = config.get("product_codes", [])
        self.items_headers = config.get("product_headers", ["#", "รหัสสินค้า", "ชื่อรายการ", "จำนวน", "ราคาต่อชิ้น", "จำนวนรวม"])
        self.supplier_names = config.get("supplierName", [])  # <-- เพิ่มบรรทัดนี้
        self.entries = {}
        self.field_vars = {}

        # Right: Fields (all possible keys from processed)เ
        self.entries = {}
        self.field_vars = {}  # เพิ่ม dict สำหรับเก็บตัวแปร checkbox

        # Document Info frame (right panel, top)
        doc_frame = tk.LabelFrame(
            self.main_frame,
            text="Document Info",
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE, "bold"),
            bg="#f5f5f5", bd=2, relief=tk.GROOVE, labelanchor="nw", padx=12, pady=10
        )
        doc_frame.grid(row=0, column=1, padx=(10,20), pady=(20,10), sticky="new")  # <-- remove columnspan
        doc_frame.grid_columnconfigure(2, weight=1)  # ปรับให้ช่องผลลัพธ์ขยาย

        for idx, (label, key) in enumerate(self.fields):
            var = tk.BooleanVar(value=True)
            self.field_vars[key] = var
            cb = tk.Checkbutton(
                doc_frame, variable=var, bg="#f5f5f5", activebackground="#f5f5f5",
                selectcolor="#f5f5f5", indicatoron=True, width=2, padx=8, pady=8,
                font=(FONT_FAMILY, DEFAULT_FONT_SIZE)
            )
            cb.grid(row=idx, column=0, sticky="w", padx=(0,10), pady=8)
            ttk.Label(doc_frame, text=label, anchor="w", width=18).grid(
                row=idx, column=1, sticky="w", padx=(0,10), pady=8
            )
            # ช่องผลลัพธ์อยู่ติดกับ label
            if key == "supplierName" and self.supplier_names:
                entry = ttk.Combobox(doc_frame, values=self.supplier_names, width=44, font=(FONT_FAMILY, DEFAULT_FONT_SIZE))
            else:
                entry = ttk.Entry(doc_frame, width=46, font=(FONT_FAMILY, DEFAULT_FONT_SIZE))
            entry.grid(row=idx, column=2, padx=(0,10), pady=8, sticky="ew")
            self.entries[key] = entry

        # สำหรับซ่อน/แสดงตารางสินค้า
        self.items_table_visible = True

        # Product List frame (right panel, bottom)
        table_frame = tk.LabelFrame(
            self.main_frame, 
            text="Product List",
            font=("Segoe UI", 13, "bold"),
            bg="#f5f5f5",
            bd=2,
            relief=tk.GROOVE,
            labelanchor="nw",
            padx=10,
            pady=8
        )
        table_frame.grid(row=1, column=1, padx=(10,20), pady=(10,10), sticky="nsew")  # <-- remove columnspan
        table_frame.grid_columnconfigure(0, weight=1)

        # --- Canvas + Scrollbar for items table ---
        self.items_table_canvas = tk.Canvas(table_frame, bg="#f5f5f5", highlightthickness=0, height=220, width=700)
        self.items_table_canvas.grid(row=0, column=0, sticky="nsew")
        self.items_table_scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.items_table_canvas.xview)
        self.items_table_scroll_x.grid(row=1, column=0, sticky="ew")
        self.items_table_canvas.configure(xscrollcommand=self.items_table_scroll_x.set)
        self.items_table_frame = tk.Frame(self.items_table_canvas, bg="#f5f5f5")
        self.items_table_window = self.items_table_canvas.create_window((0, 0), window=self.items_table_frame, anchor="nw")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # ปุ่ม Add row และ Add column ใต้ตาราง
        btns_frame = tk.Frame(table_frame, bg="#f5f5f5")
        btns_frame.grid(row=2, column=0, sticky="e", pady=(4,0))
        ttk.Button(btns_frame, text="+ Add row", command=self.add_item_row).pack(side="right", padx=(0,8))
        ttk.Button(btns_frame, text="+ Add column", command=self.add_item_column).pack(side="right", padx=(0,8))

        # ปุ่ม Submit ใต้ Add row
        ttk.Button(table_frame, text="Submit", command=self.submit).grid(row=3, column=0, sticky="e", pady=(10,2), padx=(0,8))

        # สำหรับเก็บข้อมูลตารางสินค้า
        self.items_data = []  # list of list
        self.items_editors = []  # list of Entry widgets

        # Draw initial empty preview
        self.show_empty_preview()
        self.show_items_table([])

        # Adjust grid weights for resizing
        self.main_frame.grid_rowconfigure(0, weight=0)  # Document Info row
        self.main_frame.grid_rowconfigure(1, weight=1)  # Product List row grows
        self.main_frame.grid_columnconfigure(0, weight=0)  # left panel fixed
        self.main_frame.grid_columnconfigure(1, weight=1)  # right panel flexible

    def _setup_styles(self):
        """Setup TTK styles for the application"""
        style = ttk.Style()
        style.theme_use("clam")
        
        # Configure common styles
        style.configure("TButton", 
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE), 
            padding=6
        )
        style.configure("TEntry", 
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE)
        )
        style.configure("TCombobox", 
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE)
        )
        style.configure("TCheckbutton", 
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE), 
            background="#f5f5f5"
        )
        style.configure("TLabel", 
            font=(FONT_FAMILY, DEFAULT_FONT_SIZE), 
            background="#f5f5f5"
        )
        
        # Configure Zoom button style
        style.configure("Zoom.TButton",
            font=(FONT_FAMILY, 12),
            padding=4,
            background="#f8f9fa",
            relief="flat"
        )
        style.map("TButton", 
            background=[("active", "#1ca21c")]
        )
        style.map("Zoom.TButton",
            background=[("active", "#e9ecef")],
            relief=[("pressed", "sunken")]
        )

    def _setup_main_frame(self):
        """Setup the main application frame"""
        self.main_frame = tk.Frame(self.root, background="#e9ecef")
        self.main_frame.pack(fill="both", expand=True)
        
        # Configure grid weights
        self.main_frame.grid_columnconfigure(0, weight=0)  # left panel fixed
        self.main_frame.grid_columnconfigure(1, weight=0)
        self.main_frame.grid_columnconfigure(1, weight=1)  # right panel flex

    def _setup_preview_frame(self):
        """Setup the image preview frame"""
        # self.preview_frame ถูกสร้างและ pack ใน __init__ แล้ว
        preview_container = tk.Frame(self.preview_frame, bg="#fff")
        preview_container.pack(anchor="n")  # ให้ติดบน
        self._setup_preview_canvas(preview_container)
        self._setup_zoom_controls()

    def _setup_preview_canvas(self, container):
        """Setup canvas and scrollbars for image preview"""
        canvas_container = tk.Frame(container, bg="#fff")
        canvas_container.pack(side="top")

        self.preview_canvas = tk.Canvas(
            canvas_container,
            bg="#fff",
            width=PREVIEW_WIDTH,
            height=PREVIEW_HEIGHT
        )
        self.preview_canvas.pack(side="left")

        # Add scrollbars
        self.preview_scrolly = ttk.Scrollbar(
            canvas_container,
            orient="vertical",
            command=self.preview_canvas.yview
        )
        self.preview_scrolly.pack(side="right", fill="y")

        self.preview_scrollx = ttk.Scrollbar(
            container,
            orient="horizontal",
            command=self.preview_canvas.xview
        )
        self.preview_scrollx.pack(side="bottom", fill="x")

        self.preview_canvas.configure(
            yscrollcommand=self.preview_scrolly.set,
            xscrollcommand=self.preview_scrollx.set
        )

    def _setup_zoom_controls(self):
        """Setup zoom controls for image preview"""
        zoom_frame = tk.Frame(self.preview_frame, bg="#fff", height=40)  # กำหนดความสูงคงที่
        zoom_frame.pack(side="bottom", fill="x", pady=(5,5))
        zoom_frame.pack_propagate(False)  # ป้องกันไม่ให้ frame หดตามเนื้อหา

        # จัดกลุ่มปุ่มให้อยู่กลาง
        btn_group = tk.Frame(zoom_frame, bg="#fff")
        btn_group.pack(expand=True)  # ใช้ expand=True เพื่อให้อยู่ตรงกลาง

        # สร้างปุ่มด้วย style ใหม่
        ttk.Button(
            btn_group,
            text="🔍+",
            command=self.zoom_in,
            width=6,
            style="Zoom.TButton"
        ).pack(side="left", padx=2)

        ttk.Button(
            btn_group,
            text="🔍-",
            command=self.zoom_out,
            width=6,
            style="Zoom.TButton"
        ).pack(side="left", padx=2)

        ttk.Button(
            btn_group,
            text="↺ Reset",
            command=self.reset_zoom,
            width=8,
            style="Zoom.TButton"
        ).pack(side="left", padx=2)

        # เพิ่มข้อความแสดงอัตราส่วน zoom
        self.zoom_label = tk.Label(
            btn_group,  # ย้ายไปอยู่ใน btn_group
            text="100%",
            font=(FONT_FAMILY, 12),
            bg="#fff",
            fg="#6c757d"
        )
        self.zoom_label.pack(side="left", padx=5)
        # --- เพิ่ม mouse drag scroll ---
        self.preview_canvas.bind("<ButtonPress-1>", self._start_drag)
        self.preview_canvas.bind("<B1-Motion>", self._on_drag)
        # ไม่จำเป็นต้องใช้ <ButtonRelease-1> สำหรับ scroll แบบนี้

    def _start_drag(self, event):
        self.preview_canvas.scan_mark(event.x, event.y)

    def _on_drag(self, event):
        self.preview_canvas.scan_dragto(event.x, event.y, gain=1)

    def show_empty_preview(self):
        self.preview_canvas.delete("all")
        # ใช้ค่า constants แทนการกำหนดค่าตายตัว
        self.preview_width = PREVIEW_WIDTH
        self.preview_height = PREVIEW_HEIGHT
        self.preview_canvas.config(width=self.preview_width, height=self.preview_height)
        self.preview_canvas.create_rectangle(
            0, 0, 
            self.preview_width, 
            self.preview_height, 
            fill="white", 
            outline=""
        )

    def update_nav_buttons(self):
        # Enable/disable prev/next buttons
        if not self.image_paths or len(self.image_paths) == 1:
            self.prev_btn.state(['disabled'])
            self.next_btn.state(['disabled'])
        else:
            if self.current_image_index <= 0:
                self.prev_btn.state(['disabled'])
            else:
                self.prev_btn.state(['!disabled'])
            if self.current_image_index >= len(self.image_paths) - 1:
                self.next_btn.state(['disabled'])
            else:
                self.next_btn.state(['!disabled'])

    def select_file(self):
        filetypes = [
            ("Image files", "*.png;*.jpg;*.jpeg;*"),
            ("PDF files", "*.pdf"),
        ]
        paths = filedialog.askopenfilenames(title="เลือกไฟล์ภาพหรือ PDF", filetypes=filetypes)
        if paths:
            self.image_paths = list(paths)
            self.current_image_index = 0
            self.show_image_at_index(self.current_image_index)
        else:
            self.image_paths = []
            self.current_image_index = 0
            self.show_empty_preview()
        self.update_nav_buttons()

    def show_image_at_index(self, idx):
        if not self.image_paths:
            self.show_empty_preview()
            return
        path = self.image_paths[idx]
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".pdf":
                img = self.pdf_page_to_image(path, page_num=0)
            else:
                img = Image.open(path)
            img = ImageOps.exif_transpose(img)
            self.original_img = img.copy()
            self.zoom_ratio = 1.0
            self.display_image(img)
        except Exception as e:
            messagebox.showerror("Error", f"Open image failed: {e}")
            self.show_empty_preview()
        self.update_nav_buttons()

    def show_prev_image(self):
        if self.image_paths and self.current_image_index > 0:
            self.current_image_index -= 1
            self.show_image_at_index(self.current_image_index)

    def show_next_image(self):
        if self.image_paths and self.current_image_index < len(self.image_paths) - 1:
            self.current_image_index += 1
            self.show_image_at_index(self.current_image_index)

    def display_image(self, img, zoom=False):
        # คำนวณขนาดภาพให้พอดีกับ canvas หรือขยายตาม zoom
        w, h = img.size
        if zoom:
            # ขยายตาม self.zoom_ratio
            new_size = (max(1, int(w * self.zoom_ratio)), max(1, int(h * self.zoom_ratio)))
        else:
            # ย่อให้พอดี preview panel
            ratio = min(PREVIEW_WIDTH / w, PREVIEW_HEIGHT / h)
            self.zoom_ratio = ratio  # sync zoom_ratio กับ preview
            new_size = (int(w * ratio), int(h * ratio))
        resized_img = img.resize(new_size, Image.LANCZOS)
        self.tk_img = ImageTk.PhotoImage(resized_img)
        self.preview_canvas.delete("all")
        x = (PREVIEW_WIDTH - new_size[0]) // 2
        y = (PREVIEW_HEIGHT - new_size[1]) // 2
        self.preview_canvas.create_image(x, y, anchor="nw", image=self.tk_img)
        self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all"))
        self.current_img = resized_img
        # อัปเดตข้อความแสดงอัตราส่วน zoom
        zoom_percent = int(self.zoom_ratio * 100)
        self.zoom_label.config(text=f"{zoom_percent}%")

    def zoom_in(self):
        if self.original_img is not None:
            prev_zoom = self.zoom_ratio
            self.zoom_ratio *= 1.05  # เพิ่มทีละ 5%
            percent_change = int((self.zoom_ratio / prev_zoom - 1) * 100)
            print(f"Zoom In: +{percent_change}%")
            self.display_image(self.original_img, zoom=True)

    def zoom_out(self):
        if self.original_img is not None:
            prev_zoom = self.zoom_ratio
            self.zoom_ratio /= 1.05  # ลดทีละ 5%
            percent_change = int((self.zoom_ratio / prev_zoom - 1) * 100)
            print(f"Zoom Out: {percent_change}%")
            self.display_image(self.original_img, zoom=True)

    def reset_zoom(self):
        if self.original_img is not None:
            self.zoom_ratio = min(PREVIEW_WIDTH / self.original_img.size[0], PREVIEW_HEIGHT / self.original_img.size[1])
            self.display_image(self.original_img, zoom=False)

    def toggle_items_table(self):
        self.items_table_visible = not self.items_table_visible
        if self.items_table_visible:
            self.items_table_canvas.grid()
            self.items_table_scroll_x.grid()
            self.add_row_btn.grid()
            self.submit_btn.grid()
            self.hide_btn.config(text="Hide ▲")
        else:
            self.items_table_canvas.grid_remove()
            self.items_table_scroll_x.grid_remove()
            self.add_row_btn.grid_remove()
            self.submit_btn.grid_remove()
            self.hide_btn.config(text="Show ▼")

    def show_items_table(self, items):
        # Clear previous table
        for widget in self.items_table_frame.winfo_children():
            widget.destroy()
        self.items_editors = []
        headers = self.items_headers
        col_weights = [1] + [2]*(len(headers)-1)
        # --- วาดหัวตารางแบบแก้ไขชื่อคอลัมน์ได้ ---
        self.header_editors = []
        for col, (h, w) in enumerate(zip(headers, col_weights)):
            header_entry = tk.Entry(
                self.items_table_frame, borderwidth=1, relief="solid",
                font=("Arial", 11, "bold"), justify="center", width=15
            )
            header_entry.insert(0, h)
            header_entry.grid(row=0, column=col, sticky="nsew", padx=0, pady=0, ipady=3)
            # เมื่อแก้ไขชื่อคอลัมน์แล้วอัปเดต self.items_headers
            def save_header(event, idx=col, entry=header_entry):
                self.items_headers[idx] = entry.get()
            header_entry.bind("<FocusOut>", save_header)
            header_entry.bind("<Return>", save_header)
            self.header_editors.append(header_entry)
            self.items_table_frame.grid_columnconfigure(col, weight=w, minsize=40)
        tk.Label(self.items_table_frame, text="", bg="#f5f5f5").grid(row=0, column=len(headers), sticky="nsew")
        # เตรียมข้อมูล
        # ถ้า items เป็น list of dict (จาก OCR) ให้แปลงเป็น list of list
        if items and isinstance(items[0], dict):
            self.items_data = []
            for row, item in enumerate(items, start=1):
                row_data = [
                    str(item.get('itemNo', row)),
                    item.get('itemCode', ''), 
                    item.get('itemName', ''),
                    item.get('itemUnit', ''),
                    item.get('itemUnitCost', ''),
                    item.get('itemTotalCost', '')
                ]
                while len(row_data) < len(headers):
                    row_data.append("")
                self.items_data.append(row_data)
        # ถ้า items เป็น list of list (จากการเพิ่มแถว/คอลัมน์/แก้ไข)
        elif items and isinstance(items[0], list):
            # ป้องกันกรณีคอลัมน์เพิ่ม/ลด
            self.items_data = [row[:len(headers)] + [""]*(len(headers)-len(row)) for row in items]
        # ถ้าไม่มีข้อมูลเลย ให้แสดงแถวเปล่า 1 แถว
        if not self.items_data:
            self.items_data = [[""] * len(headers)]
        # วาดตาราง
        for row_idx, row_data in enumerate(self.items_data, start=1):
            row_editors = []
            for col_idx, value in enumerate(row_data):
                # --- ถ้าเป็นคอลัมน์รหัสสินค้า ให้ใช้ Combobox ---
                if self.items_headers[col_idx] == "รหัสสินค้า":
                    cb = ttk.Combobox(
                    self.items_table_frame, values=self.product_codes,
                    font=("Arial", 11), width=13
                    )
                    cb.set(str(value) if value is not None else "")
                    cb.grid(row=row_idx, column=col_idx, sticky="nsew", padx=0, pady=0, ipady=3)
                    row_editors.append(cb)
                else:
                    e = tk.Entry(
                        self.items_table_frame, borderwidth=1, relief="solid",
                        font=("Arial", 11), justify="left", width=15
                    )
                    e.grid(row=row_idx, column=col_idx, sticky="nsew", padx=0, pady=0, ipady=3)
                    e.insert(0, value)
                    row_editors.append(e)
            # ปุ่มลบแถว (ถังขยะ)
            del_btn = tk.Button(self.items_table_frame, text="🗑", fg="#c00", relief="flat", command=partial(self.delete_item_row, row_idx-1), cursor="hand2")
            del_btn.grid(row=row_idx, column=len(row_data), sticky="nsew", padx=(2,0))
            row_editors.append(del_btn)
            self.items_editors.append(row_editors)
        # --- ปรับ scrollregion และขนาด frame ---
        self.items_table_frame.update_idletasks()
        self.items_table_canvas.config(scrollregion=self.items_table_canvas.bbox("all"))
        # กำหนดความกว้าง canvas คงที่ ไม่ขยายตามตาราง
        # (ถ้าต้องการปรับอัตโนมัติตามขนาดหน้าต่างหลัก ให้ใช้ self.items_table_canvas.config(width=self.items_table_canvas.winfo_width()))
        # ไม่ต้องปรับ width ของ canvas ตาม frame_width

    def delete_item_row(self, idx):
        if 0 <= idx < len(self.items_data):
            del self.items_data[idx]
            self.show_items_table(self.items_data)

    def add_item_row(self):
        # เพิ่มแถวใหม่ (ค่าว่าง)
        if not self.items_data:
            self.items_data = [[""] * len(self.items_headers)]
        else:
            self.items_data.append([""] * len(self.items_headers))
        self.show_items_table(self.items_data)

    def add_item_column(self):
        # เพิ่มคอลัมน์ใหม่ (ชื่อคอลัมน์อัตโนมัติ)
        new_col_name = f"คอลัมน์{len(self.items_headers)}"
        self.items_headers.append(new_col_name)
        for row in self.items_data:
            row.append("")
        self.show_items_table(self.items_data)

    def submit(self):
        messagebox.showinfo("Submit", "Submit clicked!\n(คุณสามารถนำโค้ดนี้ไปต่อยอดบันทึก/ส่งข้อมูลได้)")

    def process_ocr(self):
        if not self.file_path:
            messagebox.showwarning("No file", "กรุณาเลือกไฟล์ภาพก่อน")
            return
        try:
            ext = os.path.splitext(self.file_path)[1].lower()
            results = []
            
            if (ext == '.pdf'):
                # แปลง PDF ทุกหน้าเป็นรูปภาพ
                images = self.pdf_to_images(self.file_path)
                for i, image in enumerate(images):
                    # บันทึกเป็นไฟล์ชั่วคราว
                    temp_path = os.path.join(os.path.dirname(self.file_path), f"temp_ocr_{i}.png")
                    image.save(temp_path, "PNG")
                    
                    # ส่งไฟล์ไปยัง API
                    with open(temp_path, "rb") as f:
                        files = {'file': (os.path.basename(temp_path), f, 'image/png')}
                        headers = {'apikey': API_KEY}
                        data = {'return_image': 'false', 'return_ocr': 'false'}
                        resp = requests.post(OCR_URL, headers=headers, files=files, data=data)
                        resp.raise_for_status()
                        result = resp.json()
                        if result.get("message") == "success":
                            results.append(result)
                    
                    # ลบไฟล์ชั่วคราว
                    os.remove(temp_path)
            else:
                # กรณีเป็นรูปภาพ ทำแบบเดิม
                with open(self.file_path, "rb") as f:
                    files = {'file': (os.path.basename(self.file_path), f, 'image/png')}
                    headers = {'apikey': API_KEY}
                    data = {'return_image': 'false', 'return_ocr': 'false'}
                    resp = requests.post(OCR_URL, headers=headers, files=files, data=data)
                    resp.raise_for_status()
                    results.append(resp.json())

            # แสดงผลลัพธ์
            if results:
                # รวมข้อมูลจากทุกหน้า
                all_items = []
                for result in results:
                    if result.get("message") == "success":
                        processed = result.get("processed", {})
                        # เก็บข้อมูล fields จากหน้าแรกเท่านั้น
                        if result == results[0]:
                            for label, key in self.fields:
                                value = processed.get(key, "")
                                self.entries[key].delete(0, tk.END)
                                self.entries[key].insert(0, str(value) if value is not None else "")
                        # รวมรายการสินค้าจากทุกหน้า
                        items = processed.get("items", [])
                        all_items.extend(items)
                
                # แสดงตารางรายการสินค้าทั้งหมด
                self.show_items_table(all_items)
            else:
                messagebox.showerror("OCR Failed", "ไม่พบผลลัพธ์จาก OCR")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def load_config(self):
        json_path = os.path.join(os.path.dirname(__file__), "config_fields.json")
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                config = json.load(f)
            # รองรับไฟล์เก่า (list of list) เพื่อความเข้ากันได้
            if isinstance(config, list):
                config = {"fields": config, "product_codes": []}
            return config
        except Exception as e:
            messagebox.showerror("Error", f"โหลด config ไม่สำเร็จ: {e}")
            return {"fields": [], "product_codes": []}

    def export_excel(self):
        # เตรียมข้อมูลเฉพาะ field ที่ถูกเลือก
        data = []
        field_names = []
        field_values = []
        for label, key in self.fields:
            if self.field_vars[key].get():
                field_names.append(label)
                field_values.append(self.entries[key].get())
        # เตรียมข้อมูล items จาก Entry (self.items_editors)
        items = []
        for row_editors in self.items_editors:
            row_values = []
            for idx, e in enumerate(row_editors):
                if idx >= len(self.items_headers):
                    # ข้าม widget ที่เกินจำนวน header เช่น ปุ่มลบแถว
                    continue
                # ถ้าเป็น Combobox (รหัสสินค้า) ให้ใช้ get()
                if self.items_headers[idx] == "รหัสสินค้า" and isinstance(e, ttk.Combobox):
                    row_values.append(e.get())
                elif isinstance(e, tk.Entry):
                    row_values.append(e.get())
            if any(row_values):
                items.append(row_values)
        # --- สร้าง header และ value row สำหรับ export ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR Data"
        # รวม field headers + items headers
        export_headers = list(field_names)
        export_values = list(field_values)
        if items:
            # ต่อ header ของสินค้าทั้งหมด (ถ้ามีหลายแถว)
            for idx, row in enumerate(items, start=1):
                for col_name in self.items_headers:
                    export_headers.append(f"{col_name}")
            # ต่อค่าของสินค้าทั้งหมด
            for row in items:
                for value in row:
                    export_values.append(value)
        ws.append(export_headers)
        ws.append(export_values)
        # ปรับความกว้างคอลัมน์
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="บันทึกไฟล์ Excel"
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Export สำเร็จ", f"บันทึกไฟล์ Excel ที่\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export ผิดพลาด", str(e))

    def pdf_page_to_image(self, path, page_num=0):
        doc = fitz.open(path)
        page = doc.load_page(page_num)
        pix = page.get_pixmap()
        mode = "RGBA" if pix.alpha else "RGB"
        img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
        return img

    def pdf_to_images(self, pdf_path):
        """แปลง PDF เป็นรูปภาพทุกหน้า"""
        doc = fitz.open(pdf_path)
        images = []
        for page_num in range(doc.page_count):
            page = doc[page_num]
            pix = page.get_pixmap()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()
        return images

if __name__ == "__main__":
    root = tk.Tk()
    app = OCRApp(root)
    root.mainloop()
